from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from delivery_qc.domain.models import PackageResult, ResultCode
from delivery_qc.domain.normalize import status_matches
from delivery_qc.infrastructure.report_presentation import (
    ACTION_BY_RESULT as _ACTIONS,
    ISSUE_BY_RESULT as _ISSUES,
    STRATUS_ORDERS_URL,
    container_url as _container_url,
    package_url as _package_url,
    power_bi_filter_url as _power_bi_filter_url,
)


_INK = "1D1D1F"
_SECONDARY = "6E6E73"
_BLUE = "0071E3"
_PALE_BLUE = "EAF3FF"
_GREEN = "F0FAF2"
_DARK_GREEN = "1E7D34"
_RED = "FFF1F2"
_DARK_RED = "D70015"
_GRAY = "F5F5F7"
_WHITE = "FFFFFF"
_BORDER = Side(style="thin", color="D2D2D7")
_DISPLAY_TIMEZONE = ZoneInfo("America/New_York")


def write_excel_report(
    *,
    path: Path,
    run_id: str,
    delivery_date: date,
    created_at: datetime,
    results: tuple[PackageResult, ...],
    warehouse_history_url: str = "",
    shipping_tracking_url: str = "",
) -> None:
    workbook = Workbook()
    workbook.iso_dates = True
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    package_sheet = workbook.create_sheet("Package Review")
    container_sheet = workbook.create_sheet("Container Detail")

    _write_dashboard(
        dashboard,
        run_id,
        delivery_date,
        created_at,
        results,
        warehouse_history_url,
        shipping_tracking_url,
    )
    _write_packages(package_sheet, delivery_date, results)
    _write_containers(container_sheet, delivery_date, results)

    for sheet in workbook:
        for row in sheet:
            for cell in row:
                if isinstance(cell.value, str):
                    cell.data_type = "s"

    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = path.with_suffix(path.suffix + ".tmp")
    workbook.save(temporary_path)
    temporary_path.replace(path)


def _write_dashboard(
    sheet,
    run_id,
    delivery_date,
    created_at,
    results,
    warehouse_history_url,
    shipping_tracking_url,
) -> None:
    sheet.sheet_view.showGridLines = False
    passed = sum(result.result_code is ResultCode.PASS_COMPLETE for result in results)
    expected = sum(result.expected_count or 0 for result in results)
    outstanding = sum(result.outstanding_count for result in results)
    flagged = len(results) - passed

    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Delivery QC"
    sheet["A1"].font = Font(name="Arial", bold=True, color=_INK, size=22)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.merge_cells("A2:I2")
    sheet["A2"] = (
        f"{_format_date(delivery_date)}  ·  Updated {_format_time(created_at)}"
    )
    sheet["A2"].font = Font(name="Arial", color=_SECONDARY, size=10)
    sheet.merge_cells("A3:I3")
    sheet["A3"] = "Shadow mode  ·  Review only  ·  No emails sent"
    sheet["A3"].font = Font(name="Arial", color=_SECONDARY, italic=True, size=9)

    sheet.merge_cells("A5:I5")
    sheet.merge_cells("A6:I6")
    if flagged:
        sheet["A5"] = f"{flagged} package{'s' if flagged != 1 else ''} need attention"
        sheet["A6"] = (
            f"{outstanding} container{'s' if outstanding != 1 else ''} outstanding  ·  "
            f"{passed} of {len(results)} packages cleared"
        )
        summary_fill = _RED
        summary_color = _DARK_RED
    else:
        sheet["A5"] = "No delivery issues found"
        sheet["A6"] = (
            f"All {len(results)} scheduled package{'s are' if len(results) != 1 else ' is'} "
            "confirmed Field Received"
        )
        summary_fill = _GREEN
        summary_color = _DARK_GREEN
    for row in (5, 6):
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=summary_fill)
        sheet.cell(row, 1).alignment = Alignment(vertical="center")
    sheet["A5"].font = Font(name="Arial", bold=True, color=summary_color, size=18)
    sheet["A6"].font = Font(name="Arial", color=_INK, size=10)

    sheet["K1"] = "Run ID"
    sheet["L1"] = run_id
    sheet["K2"] = "Decision rule"
    sheet["L2"] = "PASS only when every expected container is Field Received"
    sheet["K3"] = "Expected containers"
    sheet["L3"] = expected
    sheet.column_dimensions["K"].hidden = True
    sheet.column_dimensions["L"].hidden = True

    _write_investigation_queue(
        sheet,
        results,
        warehouse_history_url,
        shipping_tracking_url,
    )

    widths = (20, 39, 18, 35, 38, 13, 13, 15, 15)
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[1].height = 34
    sheet.row_dimensions[2].height = 19
    sheet.row_dimensions[3].height = 17
    sheet.row_dimensions[4].height = 8
    sheet.row_dimensions[5].height = 30
    sheet.row_dimensions[6].height = 24
    sheet.sheet_view.zoomScale = 90


def _write_investigation_queue(
    sheet,
    results,
    warehouse_history_url,
    shipping_tracking_url,
) -> None:
    investigation_items = _investigation_items(results)
    sheet["A8"] = "Needs attention"
    sheet["A8"].font = Font(name="Arial", bold=True, color=_INK, size=14)
    sheet["I8"] = f"{len(investigation_items)} item{'s' if len(investigation_items) != 1 else ''}"
    sheet["I8"].font = Font(name="Arial", color=_SECONDARY, size=9)
    sheet["I8"].alignment = Alignment(horizontal="right")
    sheet.merge_cells("A9:I9")
    sheet["A9"] = "Open a source link to investigate. Each link is already filtered to this item."
    sheet["A9"].font = Font(name="Arial", color=_SECONDARY, size=9)
    headers = (
        "Item",
        "Package",
        "Status",
        "Why",
        "Next step",
        "Container",
        "Package",
        "Warehouse",
        "Shipping",
    )
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(10, column, header)
        cell.fill = PatternFill("solid", fgColor=_INK)
        cell.font = Font(name="Arial", bold=True, color=_WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    if not investigation_items:
        sheet.merge_cells("A11:I12")
        sheet["A11"] = "Everything scheduled is confirmed Field Received."
        sheet["A11"].fill = PatternFill("solid", fgColor=_GREEN)
        sheet["A11"].font = Font(name="Arial", bold=True, color=_DARK_GREEN, size=12)
        sheet["A11"].alignment = Alignment(horizontal="center", vertical="center")
        return

    for row_number, item in enumerate(investigation_items, start=11):
        container_url = str(item["container_url"])
        package_url = str(item["package_url"])
        warehouse_url = _power_bi_filter_url(
            warehouse_history_url,
            "Package/Name",
            str(item["package_name"]),
        )
        shipping_field = (
            "Part/Container_x002E_Name"
            if container_url
            else "Calendar/Package_x0020_Name"
        )
        shipping_value = (
            str(item["item_name"])
            if container_url
            else str(item["package_name"])
        )
        shipping_url = _power_bi_filter_url(
            shipping_tracking_url,
            shipping_field,
            shipping_value,
        )
        values = (
            item["item_name"],
            item["package_name"],
            item["status"],
            item["issue"],
            item["action"],
            "Open" if container_url else "Unavailable",
            "Open" if package_url else "Search",
            "Open" if warehouse_url else "Unavailable",
            "Open" if shipping_url else "Unavailable",
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=_BORDER)
        sheet.cell(row_number, 1).font = Font(
            name="Arial", bold=True, color=_DARK_RED, size=10
        )
        sheet.cell(row_number, 3).font = Font(name="Arial", bold=True, color=_DARK_RED)
        sheet.cell(row_number, 1).border = Border(left=Side(style="medium", color=_DARK_RED), bottom=_BORDER)
        if container_url:
            _action_link(sheet.cell(row_number, 6), container_url)
        _action_link(
            sheet.cell(row_number, 7),
            package_url or STRATUS_ORDERS_URL,
        )
        if warehouse_url:
            _action_link(sheet.cell(row_number, 8), warehouse_url)
        if shipping_url:
            _action_link(sheet.cell(row_number, 9), shipping_url)
        sheet.row_dimensions[row_number].height = 48

    queue_widths = (20, 39, 18, 35, 38, 13, 13, 15, 15)
    for column, width in enumerate(queue_widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = max(
            sheet.column_dimensions[get_column_letter(column)].width or 0,
            width,
        )
    sheet.auto_filter.ref = f"A10:I{10 + len(investigation_items)}"


def _investigation_items(results) -> list[dict[str, object]]:
    items: list[dict[str, object]] = []
    for result in results:
        project = result.delivery.project or (
            result.containers[0].project if result.containers else ""
        )
        result_items = 0
        if not result.containers:
            if result.follow_up_required:
                items.append(
                    {
                        "item_name": result.delivery.package_name,
                        "package_name": result.delivery.package_name,
                        "project": project,
                        "status": "PACKAGE NOT FOUND",
                        "issue": _ISSUES[result.result_code],
                        "action": _ACTIONS[result.result_code],
                        "container_url": "",
                        "package_url": "",
                    }
                )
            continue

        for container in result.containers:
            received = status_matches(container.status, ("Field Received",))
            if result.follow_up_required and not received:
                items.append(
                    {
                        "item_name": container.container_name,
                        "package_name": result.delivery.package_name,
                        "project": container.project or project,
                        "status": container.status,
                        "issue": _ISSUES[result.result_code],
                        "action": _ACTIONS[result.result_code],
                        "container_url": _container_url(container.container_id),
                        "package_url": _package_url(result),
                    }
                )
                result_items += 1
        if result.follow_up_required and result_items == 0:
            items.append(
                {
                    "item_name": "Package-level exception",
                    "package_name": result.delivery.package_name,
                    "project": project,
                    "status": result.result_code.value,
                    "issue": _ISSUES[result.result_code],
                    "action": _ACTIONS[result.result_code],
                    "container_url": "",
                    "package_url": _package_url(result),
                }
            )
    return items


def _external_link(cell, url: str) -> None:
    cell.hyperlink = url
    cell.style = "Hyperlink"


def _action_link(cell, url: str) -> None:
    _external_link(cell, url)
    cell.fill = PatternFill("solid", fgColor=_PALE_BLUE)
    cell.font = Font(name="Arial", bold=True, color=_BLUE, size=9, underline="single")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_packages(sheet, delivery_date, results) -> None:
    headers = (
        "Review Status",
        "Delivery Date",
        "Project",
        "Package Name",
        "Trade",
        "DL Number",
        "Expected",
        "Found",
        "Field Received",
        "Outstanding",
        "Completion %",
        "QC Result",
        "Issue",
        "Next Action",
        "Last Checked",
        "Reason Codes",
        "Warnings",
    )
    _sheet_header(
        sheet,
        len(headers),
        "Package review",
        "One row per scheduled package. Start with Needs attention rows.",
    )
    sheet.append([None] * len(headers))
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)

    for result in results:
        observed_times = sorted(
            container.observed_at for container in result.containers if container.observed_at
        )
        project = result.delivery.project or (
            result.containers[0].project if result.containers else ""
        )
        denominator = (
            result.expected_count
            if result.expected_count is not None
            else result.observed_count
        )
        completion = result.field_received_count / denominator if denominator else 0
        sheet.append(
            [
                "INVESTIGATE" if result.follow_up_required else "PASS",
                delivery_date,
                project,
                result.delivery.package_name,
                result.delivery.trade,
                result.delivery.delivery_number,
                result.expected_count,
                result.observed_count,
                result.field_received_count if result.containers else "Unverified",
                result.outstanding_count if result.containers else "Unverified",
                completion if result.containers else None,
                result.result_code.value,
                _ISSUES[result.result_code],
                _ACTIONS[result.result_code],
                _parse_datetime(observed_times[-1]) if observed_times else None,
                "; ".join(result.reason_codes),
                "; ".join(result.warnings),
            ]
        )

    _format_table_sheet(sheet, headers, len(results), "PackageReviewTable")
    for row in range(5, 5 + len(results)):
        sheet.cell(row, 2).number_format = "yyyy-mm-dd"
        for column in range(7, 11):
            sheet.cell(row, column).number_format = "#,##0"
        sheet.cell(row, 11).number_format = "0%"
        package_url = _package_url(results[row - 5])
        if package_url:
            _external_link(sheet.cell(row, 4), package_url)
    _status_formatting(sheet, 5, max(5, 4 + len(results)), 1)
    sheet.conditional_formatting.add(
        f"J5:J{max(5, 4 + len(results))}",
        CellIsRule(
            operator="greaterThan",
            formula=["0"],
            fill=PatternFill("solid", fgColor=_RED),
            font=Font(bold=True, color=_DARK_RED),
        ),
    )
    widths = (15, 13, 34, 44, 14, 12, 11, 10, 14, 12, 12, 26, 40, 44, 19, 24, 24)
    _set_widths(sheet, widths)
    for column in ("E", "L", "P", "Q"):
        sheet.column_dimensions[column].hidden = True
    sheet.sheet_properties.tabColor = _BLUE


def _write_containers(sheet, delivery_date, results) -> None:
    headers = (
        "Review Status",
        "Delivery Date",
        "Project",
        "Package Name",
        "Container",
        "Stratus Status",
        "Field Received?",
        "Last Checked",
        "Package QC Result",
        "Evidence Note",
    )
    _sheet_header(
        sheet,
        len(headers),
        "Container detail",
        "One row per matched Stratus container. Needs attention means Field Received is not confirmed.",
    )
    sheet.append([None] * len(headers))
    for column, value in enumerate(headers, start=1):
        sheet.cell(4, column, value)

    data_rows = 0
    for result in results:
        project = result.delivery.project or (
            result.containers[0].project if result.containers else ""
        )
        if not result.containers:
            sheet.append(
                [
                    "INVESTIGATE",
                    delivery_date,
                    project,
                    result.delivery.package_name,
                    "No matching container found",
                    "NOT FOUND",
                    "NO",
                    None,
                    result.result_code.value,
                    _ISSUES[result.result_code],
                ]
            )
            data_rows += 1
            continue
        for container in result.containers:
            received = status_matches(container.status, ("Field Received",))
            sheet.append(
                [
                    "PASS" if received else "INVESTIGATE",
                    delivery_date,
                    container.project or project,
                    result.delivery.package_name,
                    container.container_name,
                    container.status,
                    "YES" if received else "NO",
                    _parse_datetime(container.observed_at),
                    result.result_code.value,
                    "Container is Field Received"
                    if received
                    else "Container requires status review",
                ]
            )
            data_rows += 1

    _format_table_sheet(sheet, headers, data_rows, "ContainerDetailTable")
    for row in range(5, 5 + data_rows):
        sheet.cell(row, 2).number_format = "yyyy-mm-dd"
    last_row = max(5, 4 + data_rows)
    _status_formatting(sheet, 5, last_row, 1)
    sheet.conditional_formatting.add(
        f"G5:G{last_row}",
        FormulaRule(
            formula=['G5="NO"'],
            fill=PatternFill("solid", fgColor=_RED),
            font=Font(bold=True, color=_DARK_RED),
        ),
    )
    sheet.conditional_formatting.add(
        f"G5:G{last_row}",
        FormulaRule(
            formula=['G5="YES"'],
            fill=PatternFill("solid", fgColor=_GREEN),
            font=Font(bold=True, color=_DARK_GREEN),
        ),
    )
    widths = (15, 13, 34, 44, 22, 26, 16, 19, 26, 34)
    _set_widths(sheet, widths)
    sheet.column_dimensions["I"].hidden = True
    sheet.sheet_properties.tabColor = "8E8E93"

    row_number = 5
    for result in results:
        package_url = _package_url(result)
        if not result.containers:
            row_number += 1
            continue
        for container in result.containers:
            if package_url:
                _external_link(sheet.cell(row_number, 4), package_url)
            container_url = _container_url(container.container_id)
            if container_url:
                _external_link(sheet.cell(row_number, 5), container_url)
            row_number += 1


def _sheet_header(sheet, column_count, title, subtitle) -> None:
    last_column = get_column_letter(column_count)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    _title(sheet["A1"])
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="Arial", color=_SECONDARY, size=9)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[3].height = 8


def _format_table_sheet(sheet, headers, data_rows, table_name) -> None:
    for cell in sheet[4]:
        if cell.column <= len(headers):
            cell.fill = PatternFill("solid", fgColor=_INK)
            cell.font = Font(name="Arial", bold=True, color=_WHITE, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    if data_rows:
        table = Table(displayName=table_name, ref=f"A4:{get_column_letter(len(headers))}{4 + data_rows}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleLight9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    else:
        sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    sheet.freeze_panes = "C5"
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 18
    sheet.row_dimensions[4].height = 30
    sheet.sheet_view.zoomScale = 85
    for row in sheet.iter_rows(min_row=5, max_row=max(5, 4 + data_rows)):
        for cell in row:
            cell.font = Font(name="Arial", color=_INK, size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row[0].row].height = 30


def _status_formatting(sheet, first_row, last_row, column) -> None:
    column_letter = get_column_letter(column)
    target = f"{column_letter}{first_row}:{column_letter}{last_row}"
    sheet.conditional_formatting.add(
        target,
        FormulaRule(
            formula=[f'{column_letter}{first_row}="PASS"'],
            fill=PatternFill("solid", fgColor=_GREEN),
            font=Font(bold=True, color=_DARK_GREEN),
        ),
    )
    sheet.conditional_formatting.add(
        target,
        FormulaRule(
            formula=[f'{column_letter}{first_row}="INVESTIGATE"'],
            fill=PatternFill("solid", fgColor=_RED),
            font=Font(bold=True, color=_DARK_RED),
        ),
    )


def _set_widths(sheet, widths) -> None:
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width


def _title(cell) -> None:
    cell.font = Font(name="Arial", bold=True, color=_INK, size=20)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _parse_datetime(value: str) -> str | None:
    if not value:
        return None
    try:
        return _format_datetime(datetime.fromisoformat(value))
    except ValueError:
        return None


def _format_datetime(value: datetime) -> str:
    if value.tzinfo is not None:
        value = value.astimezone(_DISPLAY_TIMEZONE)
    return value.strftime("%Y-%m-%d %H:%M ET")


def _format_date(value: date) -> str:
    return value.strftime("%A, %B %d, %Y").replace(" 0", " ")


def _format_time(value: datetime) -> str:
    if value.tzinfo is not None:
        value = value.astimezone(_DISPLAY_TIMEZONE)
    return value.strftime("%I:%M %p ET").lstrip("0")
