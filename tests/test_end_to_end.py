from contextlib import closing
from datetime import date, datetime, timezone
import json
from pathlib import Path
import sqlite3
import tempfile
import unittest

from openpyxl import load_workbook

from delivery_qc.application.checker import run_shadow_check
from delivery_qc.config import ProjectConfig
from delivery_qc.domain.models import ContainerStatus, Delivery, PackageResult, ResultCode
from delivery_qc.infrastructure.excel_report import write_excel_report
from delivery_qc.infrastructure.html_report import write_html_report


FIXTURES = Path(__file__).parent / "fixtures"


class EndToEndTests(unittest.TestCase):
    def test_partial_run_writes_audit_report_and_unsent_draft(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            outcome = run_shadow_check(
                delivery_date=date(2026, 8, 18),
                ics_path=FIXTURES / "deliveries.ics",
                statuses_path=FIXTURES / "statuses_partial.csv",
                config=self._config(root),
            )

            self.assertEqual(ResultCode.FLAG_PARTIAL_DELIVERY, outcome.results[0].result_code)
            self.assertTrue(outcome.report_paths.markdown.exists())
            self.assertTrue(outcome.report_paths.json.exists())
            self.assertTrue(outcome.report_paths.csv.exists())
            self.assertTrue(outcome.report_paths.excel.exists())
            self.assertTrue(outcome.report_paths.html.exists())
            self.assertEqual(1, len(outcome.report_paths.drafts))
            self.assertTrue(outcome.report_paths.latest_markdown.exists())
            self.assertTrue(outcome.report_paths.latest_json.exists())
            self.assertTrue(outcome.report_paths.latest_excel.exists())
            self.assertTrue(outcome.report_paths.latest_html.exists())
            self.assertTrue(outcome.report_paths.index_html.exists())
            self.assertTrue(outcome.report_paths.history_json.exists())
            self.assertTrue(outcome.report_paths.history_csv.exists())
            self.assertTrue(outcome.report_paths.latest_drafts.exists())
            report = outcome.report_paths.markdown.read_text(encoding="utf-8")
            draft = outcome.report_paths.drafts[0].read_text(encoding="utf-8")
            self.assertIn("SHADOW MODE", report)
            self.assertIn("NO EMAILS SENT", report)
            self.assertIn("PROPOSED EMAIL — NOT SENT", draft)
            dashboard = outcome.report_paths.latest_html.read_text(encoding="utf-8")
            self.assertIn("1 package needs attention", dashboard)
            self.assertIn("Cart-2", dashboard)
            self.assertIn("Download Excel", dashboard)
            self.assertIn("latest-delivery-qc-review.xlsx", dashboard)
            history_dashboard = outcome.report_paths.index_html.read_text(encoding="utf-8")
            self.assertIn("Delivery QC history", history_dashboard)
            self.assertIn("This week", history_dashboard)
            self.assertIn("Last 7 days", history_dashboard)
            self.assertIn("Last month", history_dashboard)
            self.assertIn('id="issue-total">1', history_dashboard)
            self.assertIn("CHW-Riser-01", history_dashboard)
            self.assertIn(
                f"2026-08-18/{outcome.run_id}/delivery-qc-dashboard.html#attention-title",
                history_dashboard,
            )
            history_data = json.loads(
                outcome.report_paths.history_json.read_text(encoding="utf-8")
            )
            self.assertEqual(1, history_data["periods"]["last-7-days"]["issues"])
            self.assertEqual(1, history_data["periods"]["last-7-days"]["packages"])
            self.assertIn("delivery_date", outcome.report_paths.history_csv.read_text())

            workbook = load_workbook(outcome.report_paths.excel)
            self.assertEqual(
                ["Dashboard", "Package Review", "Container Detail"],
                workbook.sheetnames,
            )
            self.assertEqual("Delivery QC", workbook["Dashboard"]["A1"].value)
            self.assertEqual("INVESTIGATE", workbook["Package Review"]["A5"].value)
            self.assertRegex(
                workbook["Dashboard"]["A2"].value,
                r"^Tuesday, August 18, 2026  ·  Updated \d{1,2}:\d{2} [AP]M ET$",
            )
            self.assertEqual("Cart-2", workbook["Dashboard"]["A11"].value)
            self.assertEqual("Unavailable", workbook["Dashboard"]["F11"].value)
            self.assertIsNone(workbook["Dashboard"]["F11"].hyperlink)
            self.assertEqual("Search", workbook["Dashboard"]["G11"].value)
            self.assertEqual(
                "https://www.gtpstratus.com/orders",
                workbook["Dashboard"]["G11"].hyperlink.target,
            )
            workbook.close()

            with closing(sqlite3.connect(root / "state.db")) as connection:
                stored = connection.execute(
                    "SELECT mode, package_count FROM runs WHERE run_id = ?", (outcome.run_id,)
                ).fetchone()
            self.assertEqual(("shadow", 1), stored)

    def test_zero_delivery_day_still_writes_report_without_status_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            outcome = run_shadow_check(
                delivery_date=date(2026, 8, 18),
                ics_path=FIXTURES / "no_deliveries.ics",
                statuses_path=None,
                config=self._config(root),
            )

            self.assertEqual((), outcome.results)
            report = outcome.report_paths.markdown.read_text(encoding="utf-8")
            self.assertIn("calendar coverage is unverified", report)
            self.assertEqual((), outcome.report_paths.drafts)
            latest_drafts = outcome.report_paths.latest_drafts.read_text(encoding="utf-8")
            self.assertIn("No follow-up drafts were required", latest_drafts)
            dashboard = outcome.report_paths.latest_html.read_text(encoding="utf-8")
            self.assertIn("Coverage unverified", dashboard)
            history_dashboard = outcome.report_paths.index_html.read_text(encoding="utf-8")
            self.assertIn('id="issue-total">0', history_dashboard)
            self.assertIn("No reviewed dates in this range", history_dashboard)

    def test_excel_links_directly_to_stratus_package_and_container(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            path = Path(temporary_directory) / "review.xlsx"
            html_path = Path(temporary_directory) / "dashboard.html"
            project_id = "6d8ddfd2-f486-4dcf-b197-94056ab71a6c"
            model_id = "204c11ad-18cf-4939-ab27-768fa744e60e"
            package_id = "c2aaa581-4830-4c53-92d4-b068110c6827"
            container_id = "0a43a109-fb0e-4d92-a3fd-f6aed373ec58"
            result = PackageResult(
                delivery=Delivery(
                    source_uid="delivery",
                    delivery_date=date(2026, 8, 18),
                    project="Central Hospital",
                    package_name="CHW-Riser-01",
                ),
                result_code=ResultCode.FLAG_NOT_RECEIVED,
                reason_codes=("NO_FIELD_RECEIPT",),
                containers=(
                    ContainerStatus(
                        project="Central Hospital",
                        package_name="CHW-Riser-01",
                        container_name="Cart-2",
                        status="In Container",
                        observed_at="2026-08-18T18:00:00+00:00",
                        container_id=container_id,
                        package_id=package_id,
                        project_id=project_id,
                        model_id=model_id,
                    ),
                ),
                expected_count=1,
                observed_count=1,
                field_received_count=0,
                outstanding_count=1,
                follow_up_required=True,
            )
            write_excel_report(
                path=path,
                run_id="test-run",
                delivery_date=date(2026, 8, 18),
                created_at=datetime(2026, 8, 18, 20, tzinfo=timezone.utc),
                results=(result,),
                warehouse_history_url=(
                    "https://app.powerbi.com/groups/me/reports/warehouse/history"
                    "?ctid=tenant"
                ),
                shipping_tracking_url=(
                    "https://app.powerbi.com/groups/me/reports/shipping/today"
                    "?ctid=tenant"
                ),
            )
            write_html_report(
                path=html_path,
                run_id="test-run",
                delivery_date=date(2026, 8, 18),
                created_at=datetime(2026, 8, 18, 20, tzinfo=timezone.utc),
                results=(result,),
                notices=(),
                warehouse_history_url=(
                    "https://app.powerbi.com/groups/me/reports/warehouse/history"
                    "?ctid=tenant"
                ),
                shipping_tracking_url=(
                    "https://app.powerbi.com/groups/me/reports/shipping/today"
                    "?ctid=tenant"
                ),
            )

            workbook = load_workbook(path)
            package_url = (
                "https://www.gtpstratus.com/orders"
                f"?projectId={project_id}&modelId={model_id}&orderId={package_id}"
            )
            container_url = (
                "https://www.gtpstratus.com/containers"
                f"?containerId={container_id}#tab_assign"
            )
            self.assertEqual(container_url, workbook["Dashboard"]["F11"].hyperlink.target)
            self.assertEqual(package_url, workbook["Dashboard"]["G11"].hyperlink.target)
            self.assertEqual(
                "https://app.powerbi.com/groups/me/reports/warehouse/history"
                "?ctid=tenant&filter=Package%2FName%20eq%20%27CHW-Riser-01%27",
                workbook["Dashboard"]["H11"].hyperlink.target,
            )
            self.assertEqual(
                "https://app.powerbi.com/groups/me/reports/shipping/today"
                "?ctid=tenant&filter=Part%2FContainer_x002E_Name%20eq%20%27Cart-2%27",
                workbook["Dashboard"]["I11"].hyperlink.target,
            )
            self.assertEqual(package_url, workbook["Package Review"]["D5"].hyperlink.target)
            self.assertEqual(container_url, workbook["Container Detail"]["E5"].hyperlink.target)
            workbook.close()

            dashboard = html_path.read_text(encoding="utf-8")
            self.assertIn(container_url, dashboard)
            self.assertIn(package_url.replace("&", "&amp;"), dashboard)
            self.assertIn(
                "https://app.powerbi.com/groups/me/reports/warehouse/history"
                "?ctid=tenant&amp;filter=Package%2FName%20eq%20%27CHW-Riser-01%27",
                dashboard,
            )
            self.assertIn(
                "https://app.powerbi.com/groups/me/reports/shipping/today"
                "?ctid=tenant&amp;filter=Part%2FContainer_x002E_Name%20eq%20%27Cart-2%27",
                dashboard,
            )

    def _config(self, root: Path) -> ProjectConfig:
        return ProjectConfig(
            mode="shadow",
            timezone="America/New_York",
            pass_statuses=("Field Received",),
            exclusion_keywords=("meeting", "pickup", "transfer"),
            database_path=root / "state.db",
            reports_path=root / "reports",
            drafts_path=root / "drafts",
            logs_path=root / "logs",
        )


if __name__ == "__main__":
    unittest.main()
